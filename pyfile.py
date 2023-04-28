import openpyxl
from datetime import datetime, date

# Load the spreadsheet
wb = openpyxl.load_workbook('prayer_times.xlsx')
ws = wb.active

# Get today's date
now = datetime.now()
month = now.month
day = now.day

# Find the row in the spreadsheet for today's date
row = None
for r in range(2, ws.max_row+1):
    if ws.cell(row=r, column=1).value == month and ws.cell(row=r, column=2).value == day:
        row = r
        break

if row is not None:
    # Generate the HTML code with the appropriate prayer times
    fajr = ws.cell(row=row, column=3).value
    isha = ws.cell(row=row, column=8).value
    zuhr = ws.cell(row=row, column=5).value
    asr = ws.cell(row=row, column=6).value
    maghrib = ws.cell(row=row, column=7).value

    # Convert the time values to the correct format
    fajr_time = datetime.combine(date.today(), fajr).strftime('%I:%M %p')
    zuhr_time = datetime.combine(date.today(), zuhr).strftime('%I:%M %p')
    asr_time = datetime.combine(date.today(), asr).strftime('%I:%M %p')
    maghrib_time = datetime.combine(date.today(), maghrib).strftime('%I:%M %p')
    isha_time = datetime.combine(date.today(), isha).strftime('%I:%M %p')

    html = f"""
    <section class="home" id="home">
        <div class="content">
            <h3>Prayer Times</h3>
            <p id="date">{now.strftime('%B %d, %Y')}</p>
            <ul>
                <li><a href="#">Fajr<span>{fajr_time}</span></a></li>
                <li><a href="#">Zuhr<span>{zuhr_time}</span></a></li>
                <li><a href="#">Asr<span>{asr_time}</span></a></li>
                <li><a href="#">Maghrib<span>{maghrib}</span></a></li>
                <li><a href="#">Isha<span>{isha_time}</span></a></li>
            </ul>
        </div>
        <a target="_blank" href="images2/calendar.pdf" class="btn4">view calendar</a>
    </section>
    """
else:
    # If today's date is not found in the spreadsheet, show a message
    html = f"""
    <section class="home" id="home">
        <div class="content">
            <h3>Prayer Times</h3>
            <p id="date">{now.strftime('%B %d, %Y')}</p>
            <p>Prayer times for today are not available.</p>
        </div>
        <a target="_blank" href="images2/calendar.pdf" class="btn4">view calendar</a>
    </section>
    """

# Print the generated HTML code
print(html)
