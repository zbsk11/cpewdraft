import openpyxl
from datetime import datetime, date

# Load the Excel workbook
workbook = openpyxl.load_workbook('prayer_times.xlsx')
ws = workbook.active

# Set the month and day values
month = 1 # example value, replace with your own value
day = 1 # example value, replace with your own value

# Get the row number for the given month and day
row = None
for i in range(2, ws.max_row + 1):
    if ws.cell(row=i, column=1).value == month and ws.cell(row=i, column=2).value == day:
        row = i
        break

# If the row was found, read the prayer times and display them in the HTML code
if row:
    fajr_value = ws.cell(row=row, column=3).value
    zuhr_value = ws.cell(row=row, column=4).value
    asr_value = ws.cell(row=row, column=5).value
    maghrib_value = ws.cell(row=row, column=6).value
    isha_value = ws.cell(row=row, column=7).value
    
    # Convert the time values to the correct format
    fajr_time = datetime.combine(date.today(), fajr_value).strftime('%I:%M %p')
    if isinstance(zuhr_value, str):
        zuhr_time = datetime.strptime(zuhr_value, '%I:%M %p').strftime('%I:%M %p')
    else:
        zuhr_time = datetime.combine(date.today(), zuhr_value).strftime('%I:%M %p')
    if isinstance(asr_value, str):
        asr_time = datetime.strptime(asr_value, '%I:%M %p').strftime('%I:%M %p')
    else:
        asr_time = datetime.combine(date.today(), asr_value).strftime('%I:%M %p')
    if isinstance(maghrib_value, str):
        maghrib_time = datetime.strptime(maghrib_value, '%I:%M %p').strftime('%I:%M %p')
    else:
        maghrib_time = datetime.combine(date.today(), maghrib_value).strftime('%I:%M %p')
    if isinstance(isha_value, str):
        isha_time = datetime.strptime(isha_value, '%I:%M %p').strftime('%I:%M %p')
    else:
        isha_time = datetime.combine(date.today(), isha_value).strftime('%I:%M %p')

    # Display the prayer times in the HTML code
    html = '''
    <section class="home" id="home">
        <div class="content">
            <h3>Prayer Times</h3>
            <p id="date">{}/{} Prayer Times</p>
            <ul>
                <li><a href="#">Fajr<span>{}</span></a></li>
                <li><a href="#">Zuhr<span>{}</span></a></li>
                <li><a href="#">Asr<span>{}</span></a></li>
                <li><a href="#">Maghrib<span>{}</span></a></li>
                <li><a href="#">Isha<span>{}</span></a></li>
            </ul>
        </div>
        <a target="_blank" href="images2/calendar.pdf" class="btn4">view calendar</a>
    </section>
    '''.format(month, day, fajr_time, zuhr_time, asr_time, maghrib_time, isha_time)

    # Print the HTML code
    print(html)
